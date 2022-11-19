import tkinter.filedialog
from tkinter import *
from tkinter import messagebox
import os
import subprocess
import sys
import openpyxl
import pandas as pd
import shutil
# из текущего периода вычесть предыдущий


if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
elif __file__:
    application_path = os.path.dirname(__file__)
shablons_file = 'shablons.xlsx'
shablons_path = os.path.join(application_path, shablons_file)
firstFile = ''
secondFile = ''


def agregate(fileOld, fileNew, shablons_path):
    name1 = os.path.basename(fileOld).lower()
    name2 = os.path.basename(fileNew).lower()
    outputFile = fileOld + '_Разница.xlsx'
    shutil.copyfile(fileOld, outputFile)
    shablons = openpyxl.load_workbook(shablons_path)
    old_workbook = openpyxl.load_workbook(fileOld, data_only=True)
    new_workbook = openpyxl.load_workbook(fileNew, data_only=True)
    output_workbook = openpyxl.load_workbook(outputFile)
    shablons_names = shablons.sheetnames
    sheet_work_with = 0
    for num, name in enumerate(shablons_names):
        if name1.startswith(name) and name2.startswith(name):
            sheet_work_with = num
            break
    chosen_shablon_sheet = shablons[shablons_names[sheet_work_with]]
    df_shablon = pd.DataFrame(chosen_shablon_sheet.values)
    for row in df_shablon.iterrows():
        section_name = row[1][0]
        ws_name = row[1][1]
        if ws_name == 'название листа':
            continue
        upper_left = row[1][2]
        lower_left = row[1][3]
        old_sheet = old_workbook[ws_name]
        new_sheet = new_workbook[ws_name]
        output_sheet = output_workbook[ws_name]

        old_data_rows = []
        for rows in old_sheet[upper_left:lower_left]:
            data_cols = []
            for cell in rows:
                data_cols.append(cell.value)
            old_data_rows.append(data_cols)
        old_df_table = pd.DataFrame(old_data_rows)
        new_data_rows = []
        for rows in new_sheet[upper_left:lower_left]:
            data_cols = []
            for cell in rows:
                data_cols.append(cell.value)
            new_data_rows.append(data_cols)
        new_df_table = pd.DataFrame(new_data_rows)
        old_df_table.fillna(0, inplace=True)
        new_df_table.fillna(0, inplace=True)
        output_table = new_df_table-old_df_table
        for n_row, rows in enumerate(output_sheet[upper_left:lower_left]):
            for n_cell, cell in enumerate(rows):
                cell.value = output_table[n_cell][n_row]
    output_workbook.save(outputFile)
    os.startfile(outputFile)


def askFirstFile():
    global firstFile
    firstFile = tkinter.filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    print(firstFile)


def askSecondFile():
    global secondFile
    secondFile = tkinter.filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    print(secondFile)


def agregateFiles():
    global firstFile
    global secondFile
    global shablons_path
    print(firstFile, secondFile)
    if firstFile == '' or secondFile == '' or not os.path.exists(firstFile) or not os.path.exists(secondFile):
        messagebox.showwarning('але', 'сначала файлы выбери')
    else:
        agregate(firstFile, secondFile, shablons_path)


root = Tk()
root.title('Нарастающий итог')
Button(root, text='Загрузить предыдущий период', command=askFirstFile).pack(fill=X)
Button(root, text='Загрузить текущий период', command=askSecondFile).pack(fill=X)
Button(root, text='Обработать', command=agregateFiles).pack(fill=X)
Label(text="by DmitrySosnin 2022", foreground='grey').pack(fill=X)
root.mainloop()
